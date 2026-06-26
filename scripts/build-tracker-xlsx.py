#!/usr/bin/env python3
"""
Build trackers/all-sites.xlsx — one sheet per doctor site.
Columns: Date | Topic | Slug | URL | Status
Run from repo root: python scripts/build-tracker-xlsx.py
"""
import csv, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOG       = os.path.join(REPO, "page-log.csv")
PLAN_CSV  = os.path.join(REPO, "_data", "content-plan.csv")
OUT       = os.path.join(REPO, "trackers", "all-sites.xlsx")
TODAY_STR = date.today().strftime("%Y-%m-%d")

SITES = {
    "dr-owais":        ("Dr. Owais Mohammed",       "cioncancerdrowais.com"),
    "dr-vinay":        ("Dr. Vinay",                 "cioncancerdrvinay.com"),
    "dr-murali":       ("Dr. Murali",                "cioncancerdrmurali.com"),
    "dr-sandeep":      ("Dr. Sandeep",               "cioncancerdrsandeep.com"),
    "dr-kiranmayee":   ("Dr. Kiranmayee",            "cioncancerdrkiranmayee.com"),
    "dr-basudev":      ("Dr. Basudev",               "cioncancerdrbasudev.com"),
    "dr-raghvendra":   ("Dr. Raghvendra",            "cioncancerdrraghvendra.com"),
    "dr-craghavendra": ("Dr. C. Raghavendra",        "cioncancerdrcraghavendra.info"),
    "dr-imad":         ("Dr. Mohammed Imaduddin",    "cioncancerdrimad.com"),
}

# Load page-log into url->date lookup (date part only)
log_lookup = {}
if os.path.exists(LOG):
    with open(LOG, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            log_lookup[row["url"]] = row["timestamp_utc"][:10]

# Load content-plan rows grouped by doctor
plan_by_folder = {}
if os.path.exists(PLAN_CSV):
    with open(PLAN_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            folder = row["doctor_folder"]
            plan_by_folder.setdefault(folder, []).append(row)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
TODAY_FILL  = PatternFill("solid", fgColor="FFD700")
URL_FONT    = Font(color="0563C1", underline="single", size=10)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def status_style(status_str):
    """Return (Font, fill_or_None) for a status string."""
    if "Published" in status_str:
        return Font(color="00AA44", size=10), None
    elif "Today" in status_str:
        return Font(color="0000CC", size=10, bold=True), TODAY_FILL
    elif "Failed" in status_str:
        return Font(color="CC0000", size=10), None
    else:
        return Font(color="888888", size=10), None

wb = Workbook()
wb.remove(wb.active)

for folder, (doctor_name, domain) in SITES.items():
    site_dir = os.path.join(REPO, folder)

    # Slugs that appear in the plan (to avoid duplicate rows)
    plan_slugs = set()
    plan_data  = []
    for pr in plan_by_folder.get(folder, []):
        p_status = pr.get("status", "pending")
        p_slug   = pr.get("built_slug") or pr.get("slug_hint", "")
        p_url    = pr.get("built_url") or f"https://{domain}/{p_slug}.html"
        p_date   = pr.get("planned_date", "")
        p_topic  = pr.get("topic_en", "")

        if p_status == "published":
            row_status = "Published ✓"
            plan_slugs.add(p_slug)
        elif p_status == "failed":
            row_status = "Failed ✗"
            plan_slugs.add(p_slug)
        elif p_date == TODAY_STR:
            row_status = "Today →"
            plan_slugs.add(p_slug)
        else:
            row_status = "Scheduled"

        plan_data.append({
            "date":   p_date,
            "topic":  p_topic,
            "slug":   p_slug,
            "url":    p_url,
            "status": row_status,
        })

    # Pre-existing pages on disk that are NOT covered by the plan.
    # Handles two layouts:
    #   flat:   site-dir/slug.html          (8 doctors)
    #   subdir: site-dir/slug/index.html    (dr-imad PHP site)
    preexisting = []
    if os.path.isdir(site_dir):
        entries = sorted(os.listdir(site_dir))
        for entry in entries:
            entry_path = os.path.join(site_dir, entry)
            # Flat HTML file
            if entry.endswith(".html") and os.path.isfile(entry_path):
                if entry == "thank-you.html":
                    continue
                slug = entry.replace(".html", "")
                url  = f"https://{domain}/{entry}"
            # Subdirectory containing index.html (dr-imad style)
            elif os.path.isdir(entry_path):
                index = os.path.join(entry_path, "index.html")
                if not os.path.isfile(index):
                    continue
                slug = entry
                url  = f"https://{domain}/{entry}/"
            else:
                continue
            if slug in plan_slugs:
                continue
            date_str = log_lookup.get(url, log_lookup.get(url.rstrip("/"), "NA"))
            preexisting.append({
                "date":   date_str,
                "topic":  "(pre-existing page)",
                "slug":   slug,
                "url":    url,
                "status": "Published ✓",
            })

    all_rows = preexisting + plan_data
    all_rows.sort(key=lambda r: r["date"] if r["date"] and r["date"] != "NA" else "9999-99-99")

    ws = wb.create_sheet(title=folder.replace("dr-", "Dr. ").title())

    # Title row
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"{doctor_name}  ·  {domain}"
    title_cell.font  = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["Date", "Topic", "Slug", "URL", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[2].height = 20

    # Data rows
    for i, row_data in enumerate(all_rows):
        row = i + 3
        sfont, sfill = status_style(row_data["status"])
        alt = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_fill = sfill if sfill else alt

        def cell(col, val, font=None, hyperlink=None):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = row_fill
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                    vertical="center")
            c.font      = font or Font(size=10)
            if hyperlink:
                c.hyperlink = hyperlink
                c.font = URL_FONT
            return c

        cell(1, row_data["date"])
        cell(2, row_data["topic"])
        cell(3, row_data["slug"])
        u = row_data["url"]
        cell(4, u, hyperlink=u if u.startswith("http") else None)
        e = ws.cell(row=row, column=5, value=row_data["status"])
        e.font      = sfont
        e.fill      = row_fill
        e.border    = thin_border()
        e.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row].height = 16

    # Column widths: A=16, B=40, C=50, D=70, E=16
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 16

    ws.freeze_panes = "A3"

    print(f"  {folder}: {len(preexisting)} pre-existing + {len(plan_data)} planned = {len(all_rows)} rows")

os.makedirs(os.path.join(REPO, "trackers"), exist_ok=True)
wb.save(OUT)
print(f"\nSaved: {OUT}")
